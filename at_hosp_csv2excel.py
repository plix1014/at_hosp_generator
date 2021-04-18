#!/usr/bin/python3
# -*- coding: utf-8 -*-
#
# creating excel workbook from AGES and KAZ data
#   https://www.youtube.com/user/stefankrizmanich
#   https://www.youtube.com/watch?v=A_8ZNvl2ZWQ
#
#
# Author:      <plix1014@gmail.com>
#
# Created:     17.04.2021
# Copyright:   (c) 2021
# Licence:     CC BY-NC-SA http://creativecommons.org/licenses/by-nc-sa/4.0/
#-------------------------------------------------------------------------------

import ssl
import time
import os, getopt, sys
#
from datetime import date, timedelta, datetime
import requests
#
import pandas as pd

import openpyxl

from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl import load_workbook
from copy import copy


# activate debugging
INFO    = True
VERBOSE = True
DEBUG   = False
TRACE   = False

# temp dir
TMP = '/tmp'
DIR_SEP = '/'

if sys.platform == "win32":
    DIR_SEP = '\\'

# relativ path
data_home = '.' + DIR_SEP
subdir    = 'data'

# create data dir
if not os.path.exists(data_home + subdir):
    os.makedirs(data_home + subdir)


headers_agent  = {'User-Agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15'}

# web URLs
URL_DATA1 = 'https://covid19-dashboard.ages.at/data/'
URL_DATA2 = 'https://info.gesundheitsministerium.at/data/'

# http://www.kaz.bmg.gv.at/ressourcen-inanspruchnahme/betten.html
# http://www.kaz.bmg.gv.at/fileadmin/user_upload/Betten/11_T_Betten_Fachr.xlsx
URL_KAZ = 'http://www.kaz.bmg.gv.at/fileadmin/user_upload/Betten/'

# input source files
KAZ_BETTEN     = '11_T_Betten_Fachr.xlsx'
AGES_FALL      = 'CovidFallzahlen.csv'
AGES_Einwohner = 'CovidFaelle_Altersgruppe.csv'
AGES_IMPFUNG   = 'timeline-bundeslaendermeldungen.csv'

# output file
AT_HOSP     = 'AT_Hospitalisierung.xlsx'

# store Einwohner from AGES_Einwohner
BL_Einwohner = { 'Burgenland': 0,
          'Kärnten'          : 0,
          'Niederösterreich' : 0,
          'Oberösterreich'   : 0,
          'Salzburg'         : 0,
          'Steiermark'       : 0,
          'Tirol'            : 0,
          'Vorarlberg'       : 0,
          'Wien'             : 0,
          'Österreich'       : 0
          }

# e.g.: addup EW in shell for BundeslandID=10
# awk -F";" '{print $4" "$5}' CovidFaelle_Altersgruppe.csv |grep ^10| awk '{x+=$2;} END {print x}'

# --------------------------------------------------------------------

def print_dbg(level,msg):
    """ print if level is required
    """
    now = time.strftime('%a %b %d %H:%M:%S %Y LT:')
    if level:
        print(("%s %s" % (now,msg)))
    return


def check_age(file_name,age=1):
    """ check age of source files
    """

    file_mod_time = datetime.fromtimestamp(os.stat(file_name).st_mtime)  # This is a datetime.datetime object!
    now = datetime.today()
    max_delay = timedelta(days=age)
    file_age = now-file_mod_time

    if file_age > max_delay:
        print_dbg(VERBOSE,"%s date is %s. This is more than %s days ago" % (os.path.basename(file_name),file_mod_time,file_age.days))
        return 1
    else:
        return 0


def download_files():
    """ download source files
    """
    # workaround for: ssl.SSLError: [SSL: DH_KEY_TOO_SMALL] dh key too small
    requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS += 'HIGH:!DH:!aNULL'
    try:
        requests.packages.urllib3.contrib.pyopenssl.DEFAULT_SSL_CIPHER_LIST += 'HIGH:!DH:!aNULL'
    except AttributeError:
        # no pyopenssl support used / needed / available
        pass

    for fn in [AGES_FALL,AGES_Einwohner,KAZ_BETTEN, AGES_IMPFUNG]:
        fage = 1
        IS_OLD = False

        infile = data_home + subdir + DIR_SEP + fn
        if fn == AGES_IMPFUNG:
            url = URL_DATA2 + fn
        else:
            url = URL_DATA1 + fn

        if fn == KAZ_BETTEN:
            fage = 300
            url = URL_KAZ + fn

        if not os.path.isfile(infile):
            IS_OLD = True
            print_dbg(INFO,"missing file %s" % fn)

        if os.path.isfile(infile) and check_age(infile,fage):
            IS_OLD = True

        if IS_OLD:
            print_dbg(INFO,"downloading %s from %s" % (fn,url))
            r = requests.get(url=url, headers=headers_agent)
            with open(infile, 'wb') as f:
                f.write(r.content)

        else:
            print_dbg(INFO,"%s is current." % fn)




def import_ages_csv2df(fn, csv_sep=';', decsep=',', dateField='Datum'):
    """ import from csv to dataframe"""
    csv = ''

    csv1 = data_home + fn + '.csv'
    csv2 = data_home + subdir + DIR_SEP + fn
    csv3 = fn

    if os.path.isfile(csv1):
        csv = csv1
    elif os.path.isfile(csv2):
        csv = csv2
    elif os.path.isfile(csv3):
        csv = csv3
    else:
        print_dbg(INFO,"WARN - '%s' could not be read!" % csv)
        return

    print_dbg(INFO,'import csv from %s' % csv)

    df = pd.read_csv(csv, sep=csv_sep, encoding='utf-8', decimal=decsep)

    return df


def read_xlsx(xls, year=None):
    """ read xls file
    """

    print_dbg(VERBOSE,"reading %s" % xls)
    xlsin = data_home + subdir + DIR_SEP + xls

    pd_xls = pd.ExcelFile(xlsin)

    sn = len(pd_xls.sheet_names)
    print_dbg(VERBOSE,"  sheets       : %s %s" % (sn,pd_xls.sheet_names))
    print_dbg(VERBOSE,"  reading sheet: %s [%s]" % ((int(sn)-1),pd_xls.sheet_names[(int(sn)-1)]))

    df = pd.read_excel(xlsin, sheet_name = pd_xls.sheet_names[int(sn)-1], header = 3)

    # trim column names
    #df.rename(columns=lambda x: x.strip(), inplace=True)

    print_dbg(DEBUG,"Column KAZ      : %s" % df.columns)

    return df


def set_values(row, value):
    return value[row]

def run_build():
    """ build dataframes with all rows
    """

    # ages files to dataframe
    df_fa = import_ages_csv2df(AGES_FALL)
    df_ew = import_ages_csv2df(AGES_Einwohner)
    df_va = import_ages_csv2df(AGES_IMPFUNG)

    print_dbg(DEBUG,"Column FA       : %s" % df_fa.columns)
    print_dbg(DEBUG,"Column Einwohner: %s" % df_ew.columns)

    # KAZ file to dataframe
    df_bed = read_xlsx(KAZ_BETTEN);

    print_dbg(DEBUG,"")
    print_dbg(DEBUG,"-- 0 ----------------------------------")
    print_dbg(INFO,"processing source data...")

    # sum up all Einwohner
    df_ew['AnzEinwohner']     = df_ew['AnzEinwohner'].astype(int)

    for key in BL_Einwohner.keys():
        x = df_ew.loc[df_ew['Bundesland'] == key, 'AnzEinwohner'].sum()
        BL_Einwohner[key] = x


    # original column header
    # Index(['Unnamed: 0', 'Österreich', 'BGLD', 'KTN', 'NÖ', 'OÖ', 'SBG', 'STM', 'TIR', 'VLB', 'WIEN'],
    df_bed.rename(columns={'BGLD': 'Burgenland',
                           'KTN' : 'Kärnten',
                           'NÖ'  : 'Niederösterreich',
                           'OÖ'  : 'Oberösterreich',
                           'SBG' : 'Salzburg',
                           'STM' : 'Steiermark',
                           'TIR' : 'Tirol',
                           'VLB' : 'Vorarlberg',
                           'WIEN': 'Wien',
                           }, inplace=True)

    print_dbg(DEBUG,"Column Einwohner: %s" % df_bed.columns)

    # get all the ICU beds
    at_beds = {}
    for key in BL_Einwohner.keys():
        x = df_bed[key].values[0]
        at_beds[key] = x
        print_dbg(DEBUG,"Val Bed  : %s" % x)

    print_dbg(DEBUG,"Column Einwohner: %s" % df_bed.head())
    print_dbg(DEBUG,"-- 1 ----------------------------------")

    # prepare data
    # Meldedat;TestGesamt;MeldeDatum;FZHosp;FZICU;FZHospFree;FZICUFree;BundeslandID;Bundesland
    # 01.04.2020;0;01.04.2020 00:00:00;7;3;12;3;1;Burgenland
    key='Meldedat'
    df_fa[key] = pd.to_datetime(df_fa[key],format='%d.%m.%Y')
    df_fa.set_index(key, inplace=True)
    df_fa      = df_fa.sort_index()
    df_fa.fillna(0, inplace=True)

    # rename value Alle
    df_fa.loc[df_fa['Bundesland'] == 'Alle', ['Bundesland'] ] = 'Österreich'

    # calc
    df_fa['TestGesamt']        = df_fa['TestGesamt'].astype(int)
    df_fa['FZHosp']            = df_fa['FZHosp'].astype(int)
    df_fa['FZICU']             = df_fa['FZICU'].astype(int)
    df_fa['FZHospFree']        = df_fa['FZHospFree'].astype(int)
    df_fa['FZICUFree']         = df_fa['FZICUFree'].astype(int)

    df_fa['Norm. zugewiesen']  = df_fa['FZHosp'] + df_fa['FZHospFree']
    df_fa['ICU zugewiesen']    = df_fa['FZICU'] + df_fa['FZICUFree']
    df_fa['Norm. Auslastung']  = (df_fa['FZHosp'] / df_fa['Norm. zugewiesen']).map('{:.6f}'.format)
    df_fa['ICU Auslastung']    = (df_fa['FZICU'] / df_fa['ICU zugewiesen']).map('{:.6f}'.format)

    # add new column
    df_fa['ICU Betten gesamt']          = df_fa['Bundesland'].apply(set_values, args =(at_beds, ))
    df_fa['ICU Anteil f. Corona']       = (df_fa['ICU zugewiesen'] / df_fa['ICU Betten gesamt']).map('{:.6f}'.format)
    df_fa['Einwohner']                  = df_fa['Bundesland'].apply(set_values, args =(BL_Einwohner, ))
    df_fa['ICU Betten gesamt pro 100T'] = (df_fa['ICU Betten gesamt'] / df_fa['Einwohner'] * 100000).map('{:.1f}'.format)
    df_fa['v. Intensiv Total']          = (df_fa['FZICU'] / df_fa['ICU Betten gesamt']).map('{:.6f}'.format)

    # set cell type
    df_fa['ICU Betten gesamt']          = df_fa['ICU Betten gesamt'].astype(int)


    print_dbg(DEBUG,"Column FA: %s" % df_fa.columns)
    print_dbg(DEBUG,"Column FA: %s" % df_fa.head(20))

    print_dbg(DEBUG,"-- 2 -----------------------------------")

    # sheet 1: build columns for sheet 1
    #                       B            C        D            E                  F                  G       H           I                 J
    df_fa1 = df_fa.loc[:, ['TestGesamt','FZHosp','FZHospFree','Norm. zugewiesen','Norm. Auslastung','FZICU','FZICUFree','ICU zugewiesen', 'ICU Auslastung',
        'ICU Anteil f. Corona','ICU Betten gesamt','Bundesland','Einwohner', 'ICU Betten gesamt pro 100T']]
    #    K                      L                   M

    # sheet 2: build columns for sheet 2
    #                       B            C        D            E                  F                  G       H                    I           J
    df_fa2 = df_fa.loc[:, ['TestGesamt','FZHosp','FZHospFree','Norm. zugewiesen','Norm. Auslastung','FZICU','v. Intensiv Total', 'FZICUFree','ICU zugewiesen',
        'ICU Auslastung', 'ICU Anteil f. Corona','ICU Betten gesamt','Bundesland','Einwohner', 'ICU Betten gesamt pro 100T']]
    #    K                 L                      M                   N            O            P


    # Impfungen
    # Datum                    ;BundeslandID;Bevölkerung;Name      ;GemeldeteImpfungenLaender;GemeldeteImpfungenLaenderPro100
    # 2021-01-13T23:59:59+01:00;1           ;294436     ;Burgenland;496                      ;0.17
    key='Meldedat'
    df_va[key] = df_va.Datum.str.extract(pat = '([0-9]+-[0-9]+-[0-9]+)')
    df_va[key] = pd.to_datetime(df_va[key],format='%Y-%m-%d')
    df_va.set_index(key, inplace=True)
    df_va      = df_va.sort_index()
    df_va.fillna(0, inplace=True)

    df_va['Bevölkerung']                     = df_va['Bevölkerung'].astype(int)
    df_va['GemeldeteImpfungenLaender']       = df_va['GemeldeteImpfungenLaender'].astype(int)
    df_va['GemeldeteImpfungenLaenderPro100'] = df_va['GemeldeteImpfungenLaenderPro100'].astype(float)


    print_dbg(INFO," data preparation finished")

    return [ df_fa, df_fa1, df_fa2, df_va ]



def export_df(df, fn):
    """ write dataframes to multiple worksheets
    """

    fout = data_home + subdir + DIR_SEP + fn

    print_dbg(INFO,"saving data to %s" % AT_HOSP)

    # create new xls file
    writer = pd.ExcelWriter(fout, engine='openpyxl', date_format = '%Y-%m-%d')

    SN='Intensiv'
    print_dbg(VERBOSE, " write sheet name %s to file..." % SN)
    df[1].to_excel(writer, sheet_name= SN, index = True, header = True)

    SN='Total'
    print_dbg(VERBOSE, " write sheet name %s to file..." % SN)

    # last day
    df[2].tail(11).to_excel(writer, sheet_name= SN, index = True, header = True)

    # all
    #df[2].to_excel(writer, sheet_name= SN, index = True, header = True)

    SN='Impfungen'
    print_dbg(VERBOSE, " write sheet name %s to file..." % SN)
    df[3].to_excel(writer, sheet_name= SN, index = True, header = True)

    writer.save()



def format_cells(fn):
    """ apply format to float fields
        set background color for some columns
        set autofilter
        lock first row
        set column width
    """

    fout = data_home + subdir + DIR_SEP + fn

    xl_sheets   = ['Intensiv', 'Total', 'Impfungen']
    xl_sh_i_pct = [6,10,11]
    xl_sh_t_pct = [6,8,11,12]
    xl_sh_v_pct = [7]

    col_pat_list_i = [ 8,9,10,11,12]
    col_pat_list_t = [ 9,10,11,12,13,14,15,16]
    col_pat_list_v = [ 7]

    # column colors
    col_green = 'b3d09a'
    col_brick = 'ffeadc'
    col_blue  = 'e1eefa'
    col_gold  = 'fff4d2'
    col_lime  = 'e7e0f0'
    col_gray  = 'f1f1f1'

    col_pattern = [ col_green,col_brick,col_blue,col_gold,col_lime, col_gray, col_gray,col_gray,col_gray,]

    print_dbg(INFO,"formating data in xlsx")

    xl = openpyxl.load_workbook(fout)
    print_dbg(VERBOSE," xlsx    : %s" % fn)
    print_dbg(VERBOSE," sheets  : %s" % xl.sheetnames)

    for sn in xl_sheets:

        if sn == 'Intensiv':
            xl_sh_pct = xl_sh_i_pct
            col_pat_list = col_pat_list_i
        elif sn == 'Total':
            xl_sh_pct = xl_sh_t_pct
            col_pat_list = col_pat_list_t

        elif sn == 'Impfungen':
            xl_sh_pct = xl_sh_v_pct
            col_pat_list = col_pat_list_v

        ws = xl[sn]

        max_row = ws.max_row
        max_col = ws.max_column

        print_dbg(VERBOSE," ws      : %s" % ws)
        print_dbg(VERBOSE," max_row : %s" % max_row)
        print_dbg(VERBOSE," max_col : %s" % max_col)

        # format date column
        for col in ws.iter_cols(min_row=2, max_col=1, max_row= max_row):
            for cell in col:
                cell.number_format = "yyyy-mm-dd"


        # format percentage
        for n in xl_sh_pct:
            for col in ws.iter_cols(min_row=2, min_col=n,  max_col=n, max_row= max_row):
                for cell in col:
                    cell.value = float(cell.value)
                    cell.number_format = "0.00%"


        # format float
        n = max_col
        for col in ws.iter_cols(min_row=2, min_col=n,  max_col=n, max_row= max_row):
            for cell in col:
                cell.value = float(cell.value)
                cell.number_format = "#,#0.0"


        i = 0
        for n in col_pat_list:
            fill = PatternFill(patternType="solid", start_color= col_pattern[i])
            i += 1
            for col in ws.iter_cols(min_row=2, min_col=n,  max_col=n, max_row= max_row):
                for cell in col:
                    cell.fill = fill


        # set filter
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

        # set cell width
        # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        # ------------------------------------------------
        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

        ws.column_dimensions = dim_holder



    # ------------------------------------------------
    xl.save(fout)

    return


def copy_sheet_to_result(src_file,tgt_file):
    """ copy sheet from KAZ file to result file
    """

    # copy sheet 2019 to BettenFachrichtung
    fsrc = data_home + subdir + DIR_SEP + src_file
    xl_src = openpyxl.load_workbook(fsrc)

    ftgt = data_home + subdir + DIR_SEP + tgt_file
    xl_tgt = openpyxl.load_workbook(ftgt)

    return


'''
   Check whether the file exist or not.
   file_path : the input file path with name.
   Return True if the file exist otherwise return False.
   https://www.dev2qa.com/how-to-use-python-openpyxl-to-copy-excel-sheet-data-in-same-and-different-excel-file
'''
def if_file_exist(file_path):
     if os.path.exists(file_path):
         return True
     else:
         return False

'''
   Check whether the excel file contain the excel sheet.
   work_book : openpyxl.Workbook instance.
   work_sheet_name : excel sheet name.
   Return True if the excel file contain the excel sheet otherwise return False.
   https://www.dev2qa.com/how-to-use-python-openpyxl-to-copy-excel-sheet-data-in-same-and-different-excel-file
'''
def if_excel_sheet_exist(work_book, work_sheet_name):
    HAS_SHEET = False

    sheet_names_list = work_book.sheetnames
    for sheet_name in sheet_names_list:
        if sheet_name == work_sheet_name:
            HAS_SHEET = True

    return HAS_SHEET


'''
   Copy excel sheet data from one excel file to another excel file.
   source_file_path : source excel file path with name.
   source_sheet_name : source excel sheet name.
   target_file_path : target excel file path with name.

   https://www.dev2qa.com/how-to-use-python-openpyxl-to-copy-excel-sheet-data-in-same-and-different-excel-file
'''
def copy_excel_sheet_in_different_file(source_file_path, source_sheet_name, target_file_path, target_sheet_name):
    target_work_sheet = None

    # if target excel file exist then load it.
    if if_file_exist(target_file_path):
        target_work_book = load_workbook(target_file_path)
    else:
        print_dbg(ERROR,"File %s do not exist." % target_file_path,)
        return 1


    if if_file_exist(source_file_path):
        # load Workbook object source excel file.
        source_work_book = load_workbook(source_file_path)

        if if_excel_sheet_exist(source_work_book, source_sheet_name):

            print_dbg(DEBUG,"Source excel sheet %s exist." % source_sheet_name)

            # get source Worksheet object.
            source_work_sheet = source_work_book[source_sheet_name]


            # if target excel sheet exist in target excel file the return it.
            if if_excel_sheet_exist(target_work_book, target_sheet_name):
                print_dbg(VERBOSE,"sheet %s already exists" % target_sheet_name)
                #target_work_sheet = target_work_book[target_sheet_name]

            # otherwise create a new Worksheet object. 
            else:
                print_dbg(VERBOSE,"create new worksheet %s" % target_sheet_name)
                target_work_sheet = target_work_book.create_sheet(target_sheet_name)


                # loop in the source excel sheet rows.    
                row_number = 1
                for row in source_work_sheet.iter_rows():

                    # loop in the row cells.
                    cell_column_number = 1
                    for cell in row:

                        # create a target excel cell in target excel sheet.
                        target_cell = target_work_sheet.cell(row = row_number, column = cell_column_number, value = cell.value)
                        if cell.has_style:
                            #target_cell._style = copy(cell._style)

                            target_cell.font = copy(cell.font)
                            target_cell.border = copy(cell.border)
                            target_cell.fill = copy(cell.fill)
                            target_cell.number_format = copy(cell.number_format)
                            target_cell.protection = copy(cell.protection)
                            target_cell.alignment = copy(cell.alignment)

                        cell_column_number += 1

                    row_number += 1

                # set cell width
                # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
                # ------------------------------------------------
                ws = target_work_sheet
                dim_holder = DimensionHolder(worksheet=ws)

                for col in range(ws.min_column, ws.max_column + 1):
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

                ws.column_dimensions = dim_holder

                # save the target excel file.
                target_work_book.save(target_file_path)

                print_dbg(DEBUG,"Excel sheet has be copied. ")
        else:
            print_dbg(DEBUG,"Source excel sheet %s do not  exist." % source_sheet_name)

    else:
        print_dbg(DEBUG,"File %s do not exist." % source_file_path,)


# --------------------------------------------------------------------

if __name__ == "__main__":

    print_dbg(INFO,"creating excel file from AGES and KAZ data")

    # download source date if necessary
    download_files()

    df = run_build()
    export_df(df,AT_HOSP)

    source_file_path = data_home + subdir + DIR_SEP + KAZ_BETTEN
    source_sheet_name = '2019'
    target_file_path = data_home + subdir + DIR_SEP + AT_HOSP
    target_sheet_name = 'BettenFachrichtung'
    copy_excel_sheet_in_different_file(source_file_path, source_sheet_name, target_file_path, target_sheet_name)

    format_cells(AT_HOSP)
    print_dbg(INFO,"%s finished" % AT_HOSP)


